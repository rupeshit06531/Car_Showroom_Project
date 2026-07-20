from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.db.models import Q, Sum, Count
from django.utils import timezone
from django.contrib import messages
from django.conf import settings
from django.core.mail import send_mail
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required

from .forms import PurchaseForm

from django.contrib.auth.decorators import (
    login_required,
    permission_required,
)
from django.contrib.admin.views.decorators import staff_member_required

from openpyxl import Workbook
from openpyxl.styles import Font
import openpyxl

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfgen import canvas

from .models import (
    Car,
    Customer,
    Sale,
    Employee,
    Supplier,
    Purchase,
)

from .forms import (
    CarForm,
    CustomerForm,
    SaleForm,
    EmployeeForm,
    SupplierForm,
    PurchaseForm,
    ExcelUploadForm,
)

# =========================
# HOME PAGE
# =========================

@login_required
def home(request):

    # Search & Filters
    query = request.GET.get("q")
    company = request.GET.get("company")
    fuel = request.GET.get("fuel")
    price = request.GET.get("price")
    stock = request.GET.get("stock")

    cars = Car.objects.all()

    # Search
    if query:
        cars = cars.filter(
            Q(car_name__icontains=query) |
            Q(company__icontains=query) |
            Q(model__icontains=query)
        )

    # Company Filter
    if company:
        cars = cars.filter(company=company)

    # Fuel Filter
    if fuel:
        cars = cars.filter(fuel_type=fuel)

    # Price Filter
    if price == "1":
        cars = cars.filter(price__lt=500000)

    elif price == "2":
        cars = cars.filter(
            price__gte=500000,
            price__lte=1000000
        )

    elif price == "3":
        cars = cars.filter(price__gt=1000000)

    # Stock Filter
    if stock == "available":
        cars = cars.filter(stock__gt=0)

    elif stock == "out":
        cars = cars.filter(stock=0)

    # Pagination
    paginator = Paginator(cars, 10)
    page_number = request.GET.get("page")
    cars = paginator.get_page(page_number)

    recent_sales = Sale.objects.select_related(
    "customer",
    "car"
    ).order_by("-id")[:5]

    recent_customers = Customer.objects.order_by("-id")[:5]

    recent_cars = Car.objects.order_by("-id")[:5]

        # Dashboard Statistics
    total_cars = Car.objects.count()
    total_customers = Customer.objects.count()
    total_sales = Sale.objects.count()

    total_companies = Car.objects.values("company").distinct().count()

    total_value = Car.objects.aggregate(Sum("price"))["price__sum"] or 0
    total_revenue = Sale.objects.aggregate(Sum("sale_price"))["sale_price__sum"] or 0
    total_purchase = Purchase.objects.aggregate(Sum("purchase_price"))["purchase_price__sum"] or 0

    net_profit = total_revenue - total_purchase

    total_stock = Car.objects.aggregate(Sum("stock"))["stock__sum"] or 0
    low_stock = Car.objects.filter(stock__lte=2).count()

    today = timezone.now().date()

    today_sales = Sale.objects.filter(sale_date=today).count()

    today_revenue = (
        Sale.objects.filter(sale_date=today)
        .aggregate(Sum("sale_price"))["sale_price__sum"] or 0
    )

    companies = Car.objects.values_list("company", flat=True).distinct()
    fuel_types = Car.objects.values_list("fuel_type", flat=True).distinct()

    months = []
    sales_count = []
    revenue = []

    current_month = timezone.now().month
    current_year = timezone.now().year

    for i in range(11, -1, -1):
        month = current_month - i
        year = current_year

        if month <= 0:
            month += 12
            year -= 1

        months.append(f"{month}/{year}")

        sales_count.append(
            Sale.objects.filter(
                sale_date__year=year,
                sale_date__month=month
            ).count()
        )

        revenue.append(
            Sale.objects.filter(
                sale_date__year=year,
                sale_date__month=month
            ).aggregate(Sum("sale_price"))["sale_price__sum"] or 0
        )

    company_names = []
    company_sales = []

    for company in companies:
        company_names.append(company)
        company_sales.append(Sale.objects.filter(car__company=company).count())

    top_cars = (
        Sale.objects.values("car__car_name")
        .annotate(total_sold=Count("id"))
        .order_by("-total_sold")[:5]
    )

    best_employee = (
        Sale.objects.values("employee__name")
        .annotate(total_sales=Count("id"))
        .order_by("-total_sales")
        .first()
    )

    return render(
        request,
        "showroom/home.html",
        {
            "cars": cars,
            "total_cars": total_cars,
            "total_customers": total_customers,
            "total_sales": total_sales,
            "total_companies": total_companies,
            "total_value": total_value,
            "total_revenue": total_revenue,
            "total_purchase": total_purchase,
            "net_profit": net_profit,
            "total_stock": total_stock,
            "low_stock": low_stock,
            "today_sales": today_sales,
            "today_revenue": today_revenue,
            "companies": companies,
            "fuel_types": fuel_types,
            "months": months,
            "sales_count": sales_count,
            "revenue": revenue,
            "company_names": company_names,
            "company_sales": company_sales,
            "top_cars": top_cars,
            "best_employee": best_employee,
            "recent_sales": recent_sales,
            "recent_customers": recent_customers,
            "recent_cars": recent_cars,
        },
)


# =========================
# ADD CAR
# =========================

@login_required
@permission_required("showroom.add_car", raise_exception=True)
def add_car(request):

    if request.method == "POST":

        form = CarForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Car added successfully."
            )

            return redirect("home")

    else:

        form = CarForm()

    return render(
        request,
        "showroom/add_car.html",
        {
            "form": form
        }
    )

# =========================
# EDIT CAR
# =========================

@login_required
@permission_required("showroom.change_car", raise_exception=True)
def edit_car(request, id):

    car = get_object_or_404(
        Car,
        id=id
    )

    if request.method == "POST":

        form = CarForm(
            request.POST,
            request.FILES,
            instance=car
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Car updated successfully."
            )

            return redirect("home")

    else:

        form = CarForm(instance=car)

    return render(
        request,
        "showroom/edit_car.html",
        {
            "form": form,
            "car": car,
        }
    )

# =========================
# DELETE CAR
# =========================

@login_required
@permission_required("showroom.delete_car", raise_exception=True)
def delete_car(request, id):

    car = get_object_or_404(
        Car,
        id=id
    )

    if request.method == "POST":

        car.delete()

        messages.success(
            request,
            "Car deleted successfully."
        )

        return redirect("home")

    return render(
        request,
        "showroom/delete_car.html",
        {
            "car": car
        }
    )

# =========================
# CAR DETAIL
# =========================

@login_required
def car_detail(request, id):

    car = get_object_or_404(
        Car,
        id=id
    )

    return render(
        request,
        "showroom/car_detail.html",
        {
            "car": car
        }
    )

# =========================
# EXPORT CARS TO EXCEL
# =========================

@login_required
def export_cars_excel(request):

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Cars"

    headers = [
        "ID",
        "Car Name",
        "Company",
        "Model",
        "Color",
        "Fuel",
        "Price",
        "Stock",
    ]

    for col, header in enumerate(headers, start=1):

        cell = sheet.cell(row=1, column=col)

        cell.value = header
        cell.font = Font(bold=True)

    for car in Car.objects.all():

        sheet.append([
            car.id,
            car.car_name,
            car.company,
            car.model,
            car.color,
            car.fuel_type,
            car.price,
            car.stock,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="cars.xlsx"'

    workbook.save(response)

    return response


# =========================
# CAR LIST
# =========================

@login_required
def car_list(request):

    cars = Car.objects.all().order_by("-id")

    query = request.GET.get("q")

    if query:
        cars = cars.filter(
            Q(car_name__icontains=query) |
            Q(company__icontains=query) |
            Q(model__icontains=query)
        )

    paginator = Paginator(cars, 10)
    page = request.GET.get("page")
    cars = paginator.get_page(page)

    return render(
    request,
    "showroom/car_list.html",
    {
        "cars": cars,
        "total_cars": Car.objects.count(),
        "in_stock": Car.objects.filter(stock__gt=0).count(),
        "out_stock": Car.objects.filter(stock=0).count(),
        "total_companies": Car.objects.values("company").distinct().count(),
    }
)



# =========================
# CUSTOMER LIST
# =========================

@login_required
def customer_list(request):

    customers = Customer.objects.all().order_by("-id")

    query = request.GET.get("q")

    if query:
        customers = customers.filter(
            Q(name__icontains=query) |
            Q(mobile__icontains=query) |
            Q(email__icontains=query)
        )

    paginator = Paginator(customers, 10)
    page_number = request.GET.get("page")
    customers = paginator.get_page(page_number)

    return render(
        request,
        "showroom/customer_list.html",
        {
    "customers": customers,
    "total_customers": Customer.objects.count(),
}
    )

# =========================
# ADD CUSTOMER
# =========================

@login_required
def add_customer(request):

    if request.method == "POST":

        form = CustomerForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Customer added successfully."
            )

            return redirect("customer_list")

    else:

        form = CustomerForm()

    return render(
        request,
        "showroom/add_customer.html",
        {
            "form": form
        }
    )

# =========================
# EDIT CUSTOMER
# =========================

@login_required
def edit_customer(request, id):

    customer = get_object_or_404(
        Customer,
        id=id
    )

    if request.method == "POST":

        form = CustomerForm(
            request.POST,
            instance=customer
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Customer updated successfully."
            )

            return redirect("customer_list")

    else:

        form = CustomerForm(instance=customer)

    return render(
        request,
        "showroom/edit_customer.html",
        {
            "form": form,
            "customer": customer
        }
    )


# =========================
# DELETE CUSTOMER
# =========================

@login_required
def delete_customer(request, id):

    customer = get_object_or_404(
        Customer,
        id=id
    )

    if request.method == "POST":

        customer.delete()

        messages.success(
            request,
            "Customer deleted successfully."
        )

        return redirect("customer_list")

    return render(
        request,
        "showroom/delete_customer.html",
        {
            "customer": customer
        }
    )

# =========================
# CUSTOMER HISTORY
# =========================

@login_required
def customer_history(request, id):

    customer = get_object_or_404(
        Customer,
        id=id
    )

    sales = (
        Sale.objects
        .select_related(
            "car",
            "employee"
        )
        .filter(customer=customer)
        .order_by("-sale_date")
    )

    total_amount = (
        sales.aggregate(
            Sum("sale_price")
        )["sale_price__sum"] or 0
    )

    return render(
        request,
        "showroom/customer_history.html",
        {
            "customer": customer,
            "sales": sales,
            "total_amount": total_amount,
        }
    )

# =========================
# SALES LIST
# =========================

@login_required
def sale_list(request):

    sales = (
        Sale.objects
        .select_related(
            "customer",
            "car",
            "employee"
        )
        .order_by("-sale_date")
    )

    query = request.GET.get("q")
    employee = request.GET.get("employee")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if query:

        sales = sales.filter(

            Q(invoice_number__icontains=query) |

            Q(customer__name__icontains=query) |

            Q(car__car_name__icontains=query)

        )

    if employee:
        sales = sales.filter(employee_id=employee)

    if start_date and end_date:

        sales = sales.filter(
            sale_date__range=[start_date, end_date]
        )

    employees = Employee.objects.all().order_by("name")

    paginator = Paginator(sales, 10)

    page = request.GET.get("page")

    sales = paginator.get_page(page)

    return render(
        request,
        "showroom/sale_list.html",
        {
            "sales": sales,
            "employees": employees,
        },
    )


@login_required
def edit_sale(request, id):

    sale = get_object_or_404(Sale, id=id)

    if request.method == "POST":

        form = SaleForm(
            request.POST,
            instance=sale
        )

        if form.is_valid():
            form.save()
            messages.success(request, "Sale updated successfully.")
            return redirect("sale_list")

    else:
        form = SaleForm(instance=sale)

    return render(
        request,
        "showroom/edit_sale.html",
        {
            "form": form,
            "sale": sale,
        }
    )


@login_required
def delete_sale(request, id):

    sale = get_object_or_404(Sale, id=id)

    if request.method == "POST":
        sale.delete()
        messages.success(request, "Sale deleted successfully.")
        return redirect("sale_list")

    return render(
        request,
        "showroom/delete_sale.html",
        {
            "sale": sale
        }
    )


# =========================
# ADD SALE
# =========================

@login_required
def add_sale(request):

    if request.method == "POST":

        form = SaleForm(request.POST)

        if form.is_valid():

            sale = form.save()

            if sale.customer.email:

                send_mail(

                    subject="Car Purchase Invoice",

                    message=f"""
Hello {sale.customer.name},

Thank you for purchasing {sale.car.car_name}.

Invoice Number : {sale.invoice_number}

Amount : ₹{sale.sale_price}

Purchase Date : {sale.sale_date}

Thank you for choosing our showroom.
""",

                    from_email=settings.DEFAULT_FROM_EMAIL,

                    recipient_list=[sale.customer.email],

                    fail_silently=True,

                )

            messages.success(
                request,
                "Sale added successfully."
            )

            return redirect("sale_list")

    else:

        form = SaleForm()

    return render(
        request,
        "showroom/add_sale.html",
        {
            "form": form
        }
    )


# =========================
# INVOICE
# =========================

@login_required
def invoice(request, id):

    sale = get_object_or_404(
        Sale,
        id=id
    )

    return render(
        request,
        "showroom/invoice.html",
        {
            "sale": sale
        }
    )


# =========================
# DOWNLOAD INVOICE PDF
# =========================

@login_required
def download_invoice(request, id):

    sale = get_object_or_404(
        Sale,
        id=id
    )

    response = HttpResponse(
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        f'attachment; filename="Invoice_{sale.invoice_number}.pdf"'
    )

    document = SimpleDocTemplate(response)

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph(
            "<b>CAR SHOWROOM MANAGEMENT SYSTEM</b>",
            styles["Title"]
        )
    )

    elements.append(
        Paragraph(
            "Customer Invoice",
            styles["Heading2"]
        )
    )

    table_data = [

        ["Invoice No", sale.invoice_number],

        ["Customer", sale.customer.name],

        ["Car", sale.car.car_name],

        ["Company", sale.car.company],

        ["Model", sale.car.model],

        ["Sale Price", f"₹ {sale.sale_price}"],

        ["Sale Date", str(sale.sale_date)],

    ]

    table = Table(
        table_data,
        colWidths=[170, 250]
    )

    table.setStyle(
        TableStyle([

            ("GRID", (0,0), (-1,-1), 1, colors.black),

            ("BACKGROUND", (0,0), (0,-1), colors.lightgrey),

            ("BOTTOMPADDING", (0,0), (-1,-1), 8),

            ("FONTNAME", (0,0), (-1,-1), "Helvetica"),

        ])
    )

    elements.append(table)

    document.build(elements)

    return response

    
# =========================
# EMPLOYEE LIST
# =========================

@login_required
def employee_list(request):

    employees = Employee.objects.all().order_by("name")

    query = request.GET.get("q")

    if query:
        employees = employees.filter(
            Q(name__icontains=query)
        )

    paginator = Paginator(employees, 10)
    page = request.GET.get("page")
    employees = paginator.get_page(page)

    return render(
        request,
        "showroom/employee_list.html",
        {
            "employees": employees
        }
    )


# =========================
# ADD EMPLOYEE
# =========================

@login_required
@staff_member_required
def add_employee(request):

    if request.method == "POST":

        form = EmployeeForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Employee added successfully."
            )

            return redirect("employee_list")

    else:

        form = EmployeeForm()

    return render(
        request,
        "showroom/add_employee.html",
        {
            "form": form
        }
    )


# =========================
# EDIT EMPLOYEE
# =========================

@login_required
@staff_member_required
def edit_employee(request, id):

    employee = get_object_or_404(Employee, id=id)

    if request.method == "POST":

        form = EmployeeForm(
            request.POST,
            request.FILES,
            instance=employee
        )

        if form.is_valid():
            form.save()
            messages.success(request, "Employee updated successfully.")
            return redirect("employee_list")

    else:
        form = EmployeeForm(instance=employee)

    return render(
    request,
    "showroom/edit_employee.html",
    {
        "form": form,
        "employee": employee,
    }
)

# =========================
# DELETE EMPLOYEE
# =========================

@login_required
@staff_member_required
def delete_employee(request, id):

    employee = get_object_or_404(Employee, id=id)

    if request.method == "POST":
        employee.delete()
        messages.success(request, "Employee deleted successfully.")
        return redirect("employee_list")

    return render(
        request,
        "showroom/delete_employee.html",
        {
            "employee": employee
        }
    )


# =========================
# SUPPLIER LIST
# =========================

@login_required
def supplier_list(request):

    suppliers = Supplier.objects.all().order_by("-id")

    return render(
        request,
        "showroom/supplier_list.html",
        {
            "suppliers": suppliers
        }
    )


# =========================
# ADD SUPPLIER
# =========================

@login_required
def add_supplier(request):

    if request.method == "POST":

        form = SupplierForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, "Supplier added successfully.")
            return redirect("supplier_list")

    else:
        form = SupplierForm()

    return render(
        request,
        "showroom/add_supplier.html",
        {
            "form": form
        }
    )


# =========================
# EDIT SUPPLIER
# =========================

@login_required
def edit_supplier(request, id):

    supplier = get_object_or_404(Supplier, id=id)

    if request.method == "POST":

        form = SupplierForm(request.POST, instance=supplier)

        if form.is_valid():
            form.save()
            messages.success(request, "Supplier updated successfully.")
            return redirect("supplier_list")

    else:
        form = SupplierForm(instance=supplier)

    return render(
        request,
        "showroom/edit_supplier.html",
        {
            "form": form,
            "supplier": supplier,
        }
    )


# =========================
# DELETE SUPPLIER
# =========================

@login_required
def delete_supplier(request, id):

    supplier = get_object_or_404(Supplier, id=id)

    if request.method == "POST":
        supplier.delete()
        messages.success(request, "Supplier deleted successfully.")
        return redirect("supplier_list")

    return render(
        request,
        "showroom/delete_supplier.html",
        {
            "supplier": supplier
        }
    )



from django.shortcuts import render
from .models import Sale


@login_required
def sales_report(request):

    sales = Sale.objects.select_related(
        "customer",
        "car",
        "employee"
    ).all()

    return render(
        request,
        "showroom/sales_report.html",
        {
            "sales": sales
        }
    )


@login_required
def purchase_list(request):
    purchases = Purchase.objects.select_related("supplier").all().order_by("-id")

    return render(
        request,
        "showroom/purchase_list.html",
        {
            "purchases": purchases
        }
    )

@login_required
def add_purchase(request):
    if request.method == "POST":
        form = PurchaseForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("purchase_list")
    else:
        form = PurchaseForm()

    return render(
        request,
        "showroom/add_purchase.html",
        {
            "form": form
        }
    )

@login_required
def edit_purchase(request, id):
    purchase = get_object_or_404(Purchase, id=id)

    if request.method == "POST":
        form = PurchaseForm(request.POST, instance=purchase)
        if form.is_valid():
            form.save()
            return redirect("purchase_list")
    else:
        form = PurchaseForm(instance=purchase)

    return render(
        request,
        "showroom/add_purchase.html",
        {
            "form": form
        }
    )


@login_required
def invoice_pdf(request, sale_id):
    sale = get_object_or_404(Sale, id=sale_id)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="invoice_{sale.invoice_number}.pdf"'

    p = canvas.Canvas(response)

    p.setFont("Helvetica-Bold", 18)
    p.drawString(180, 800, "CAR SHOWROOM INVOICE")

    p.setFont("Helvetica", 12)
    p.drawString(50, 760, f"Invoice No : {sale.invoice_number}")
    p.drawString(50, 735, f"Customer   : {sale.customer.name}")
    p.drawString(50, 710, f"Car        : {sale.car.car_name}")
    p.drawString(50, 685, f"Employee   : {sale.employee.name}")
    p.drawString(50, 660, f"Price      : ₹{sale.sale_price}")
    p.drawString(50, 635, f"Date       : {sale.sale_date}")

    p.showPage()
    p.save()

    return response




@login_required
def sales_report_pdf(request):

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="sales_report.pdf"'

    p = canvas.Canvas(response)

    p.setFont("Helvetica-Bold", 18)
    p.drawString(180, 800, "Car Showroom Sales Report")

    p.setFont("Helvetica", 11)

    y = 760

    sales = Sale.objects.select_related(
        "customer",
        "car",
        "employee"
    )

    for sale in sales:

        p.drawString(
            40,
            y,
            f"{sale.invoice_number} | {sale.customer.name} | {sale.car.car_name} | ₹{sale.sale_price}"
        )

        y -= 20

        if y < 60:
            p.showPage()
            p.setFont("Helvetica", 11)
            y = 800

    p.save()

    return response

@login_required
def delete_purchase(request, id):
    purchase = get_object_or_404(Purchase, id=id)

    if request.method == "POST":
        purchase.delete()
        return redirect("purchase_list")

    return render(request, "showroom/delete_purchase.html", {
        "purchase": purchase
    })

@login_required
def import_cars(request):
    if request.method == "POST":
        form = ExcelUploadForm(request.POST, request.FILES)

        if form.is_valid():
            return redirect("home")
    else:
        form = ExcelUploadForm()

    return render(
        request,
        "showroom/import_cars.html",
        {
            "form": form
        }
    )

@login_required
def export_sales_excel(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    headers = [
        "Invoice",
        "Customer",
        "Car",
        "Employee",
        "Sale Price",
        "Sale Date",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)

    row = 2

    for sale in Sale.objects.select_related(
        "customer",
        "car",
        "employee"
    ):

        ws.cell(row=row, column=1).value = sale.invoice_number
        ws.cell(row=row, column=2).value = sale.customer.name
        ws.cell(row=row, column=3).value = sale.car.car_name
        ws.cell(row=row, column=4).value = sale.employee.name if sale.employee else ""
        ws.cell(row=row, column=5).value = sale.sale_price
        ws.cell(row=row, column=6).value = str(sale.sale_date)

        row += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = 'attachment; filename="sales_report.xlsx"'

    wb.save(response)

    return response

